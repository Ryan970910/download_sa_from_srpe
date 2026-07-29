"""Read the list of dev_ids from the source Excel file."""
from __future__ import annotations

from typing import List

import openpyxl

from config import EXCEL_FILE


def read_dev_ids(path=EXCEL_FILE) -> List[str]:
    """Return all non-empty dev_id values from column A (skipping the header)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    dev_ids: List[str] = []
    for index, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:  # completely empty row
            continue
        value = row[0]
        if index == 0:
            # Header row. Skip it but be lenient: if it is actually a numeric
            # id (rare), still treat it as data.
            if value is None or str(value).strip().lower() in {"dev_id", "devid", "dev id", "id"}:
                continue
        if value is None:
            continue
        text = str(value).strip()
        if text:
            dev_ids.append(text)
    return dev_ids


if __name__ == "__main__":
    ids = read_dev_ids()
    print(f"Total dev_ids: {len(ids)}")
    print("First 10:", ids[:10])
    print("Last 5 :", ids[-5:])
    print("Unique :", len(set(ids)))
